import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from copy import copy

#border styles
thin = Side(border_style="thin", color="000000")
medium = Side(border_style="medium", color="000000")

def style_range(sheet, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param sheet:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = sheet[cell_range.split(":")[0]]
    if alignment:
        sheet.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = sheet[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill

def insert_rows_withStyle(sheet, row_index, amount, copy_style):
    """
    Insert row or rows before row == row_index; Copy the styles of row_index to inserted rows

    :param sheet:  Excel worksheet instance
    :param row_index: Row index to start inserting
    :param amount: number of rows will be inserted
    :param copy_style: boolean
    """

    original_cell = []
    for i in range(1, sheet.max_column):  
        original_cell.append(sheet.cell(row_index, i))

    sheet.insert_rows(row_index, amount)
    if copy_style:
        for row in range(row_index, row_index + amount + 1):
            for i in range(1, sheet.max_column):
                sheet.cell(row, i).font = copy(original_cell[i - 1].font)
                sheet.cell(row, i).border = copy(original_cell[i - 1].border)
                sheet.cell(row, i).fill = copy(original_cell[i - 1].fill)
                sheet.cell(row, i).number_format = copy(original_cell[i - 1].number_format)
                sheet.cell(row, i).protection = copy(original_cell[i - 1].protection)
                sheet.cell(row, i).alignment = copy(original_cell[i - 1].alignment)

if __name__ == "__main__":
    wb = load_workbook('temp.xlsx')
    sheet = wb.active
    insert_rows_withStyle(sheet, 1, 5, True)
    # Apply for merged cells
    border = Border(top = medium, left = medium, right = thin, bottom = thin)
    style_range(sheet, 'A1:B20', border = border)
